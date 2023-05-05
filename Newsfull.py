import requests
import pandas as pd
import datetime
from bs4 import BeautifulSoup
from requests.exceptions import ConnectionError
from openpyxl import load_workbook, Workbook
import datetime
import os
import random
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment


#section 1 for digitimes
def login_search(keywords):
    print(keywords)
    # Set up session to persist cookies across requests
    session = requests.Session()

    # Log in to the website
    login_url = 'https://www.digitimes.com/newregister/join.asp'
    login_data = {
        'username': 'berita.pasar@qdos.com.my',
        'password': 'Qdos'
    }
    response = session.post(login_url, data=login_data)

    # Check if login was successful
    if response.status_code == 200:
        login_status =1
        # print('Login successful')
    else:
        login_status =0
        # print('Login failed')

    search_url = 'https://www.digitimes.com/search/results.asp'
    search_params = {'searchtype': '1', 'q': keywords}

    response = requests.get(search_url, params=search_params)
    html_content = response.text

    soup = BeautifulSoup(html_content, 'html.parser')

    post_full_panel = soup.find("div", {"id": "result"})
    post_full_row = post_full_panel.find_all("div", {"class": "col-12 pm-0"})

    post_full_date = []
    post_full_link = []
    post_full_title = []

    for post_full_R in post_full_row:
        
        post_a = post_full_R.find('a')
        post_date = post_full_R.find(
                'div', {'class': 'date'})
        if post_a is not None:
            post_a_append = 'https://www.digitimes.com/' + post_a.attrs['href']
            post_text = post_a.text
            # get link and title
            post_full_link.append(post_a_append)
            post_full_title.append(post_text)

        else:
            post_date = post_full_R.find(
                'div', {'class': 'date'}).text
            puredate = datetime.datetime.strptime(post_date, '%A %d %B %Y').date()
            # get date
            post_full_date.append(puredate)

    # print(len(post_full_title))
    # print(len(post_full_link))
    # print(len(post_full_date))
    return post_full_date, post_full_link, post_full_title 

def main_sub(inputs):
    #ADD PARAMETER TO PASS
    lenght=inputs   
    today_date = datetime.datetime.now().strftime("%Y-%m-%d")

    #create excel file
    username = os.getlogin()
    wb = Workbook() 
    filename = "WebsScraping-News"+datetime.date.today().strftime("%Y%m%d")+"-"+str(random.randint(0, 10))+".xlsx"
    filepath = r"C:\\Users\\"+username+"\\Downloads\\" + filename +""
    wb.save(filepath)

    #create excel sheets
    ws1 = wb.create_sheet("Sheet_A")
    ws1.title = "News"
    ws2 = wb.create_sheet("Sheet_B")
    ws2.title = "Digitimes-" + lenght[0]
    ws3 = wb.create_sheet("Sheet_C")
    ws3.title = "Digitimes-" + lenght[1]
    ws4 = wb.create_sheet("Sheet_D")
    ws4.title = "Digitimes-" + lenght[2]
    wb.save(filepath)

    #load excel file
    wb = load_workbook(filepath) 

    #delete excel sheets
    del wb['Sheet']
    
    # Loop to insert data in a specific worksheet
    for value in lenght:

        post_full_date,post_full_link,post_full_title = login_search(value)

        # Select sheet
        sheet = wb["Digitimes-"+value]

        #add color background
        sheet['A3'].fill = PatternFill(start_color="6869ee", end_color="6869ee", fill_type = "solid")
        sheet['B3'].fill = PatternFill(start_color="6869ee", end_color="6869ee", fill_type = "solid")
        sheet['C3'].fill = PatternFill(start_color="6869ee", end_color="6869ee", fill_type = "solid")

        # Assign Data
        sheet['A1'] = 'Excel creation date : ' + today_date
        sheet['B1']= 'Digitimes Crawling'
        #
        sheet['A3'] = 'Title'
        sheet['A3'].alignment = Alignment(horizontal='center')
        sheet['B3'] = 'Post Date'
        sheet['B3'].alignment = Alignment(horizontal='center')
        sheet['C3'] = 'Link'
        sheet['C3'].alignment = Alignment(horizontal='center')

        #bold text
        sheet['B1'].font = Font(bold=True, size=20)
        sheet['A3'].font = Font(bold=True)
        sheet['B3'].font = Font(bold=True)
        sheet['C3'].font = Font(bold=True)

        #coloums width
        sheet.column_dimensions['A'].width = 100
        sheet.column_dimensions['B'].width = 20    
        sheet.column_dimensions['C'].width = 30

        #LOOP BY ADD 2 COLUMNS
        #------------------------------------------------------------company1
        for i in range(len(post_full_date)):
            if post_full_date[i] > datetime.date.today()-datetime.timedelta(days=30*3):
                # print("a")
                sheet.cell(row=4+i, column=1).value = post_full_title[i]
                sheet.cell(row=4+i, column=2).value = post_full_date[i]
                sheet.cell(row=4+i, column=2).alignment = Alignment(horizontal='center')
                sheet.cell(row=4+i, column=3).value = '=HYPERLINK("{}", "{}")'.format(post_full_link[i], "Link")
                sheet.cell(row=4+i, column=3).font = Font(color='6869ee')
                sheet.cell(row=4+i, column=3).alignment = Alignment(horizontal='center')

    news_main(wb,today_date)
    wb.save(filepath)

# section 2 for news
def news():
    url = 'https://www.semiconductors.org/policies/tax/market-data/?type=post'

    response = requests.get(url, headers={
        'User-Agent':  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'})
    html = response.text

    soup = BeautifulSoup(html, 'html.parser')
    
    # for loop
    post_full = soup.find_all('div', {'class': 'col-sm-8'})
    post_full_date = []
    post_full_link = []
    post_full_title = []

    for post_full_1 in post_full:

        postdate = post_full_1.find(
            'div', {'class': 'resource-item-meta'}).text
        postdate1 = postdate.split(":")
        puredate = datetime.datetime.strptime(
            postdate1[1].strip(), "%m/%d/%y").date()
        alink = post_full_1.find('a').attrs['href']
        posttiltle = post_full_1.find('h3').text

        if puredate > datetime.date.today()-datetime.timedelta(days=30*6):
            post_full_date.append(puredate)
            post_full_link.append(alink)
            post_full_title.append(posttiltle)

    return post_full_date, post_full_link, post_full_title 

def news_main(wb,today_date):
    # print(wb)
    # print(today_date)
    #ADD PARAMETER TO PASS
    #get value from post_full_date
    post_full_date,post_full_link,post_full_title = news()

    # Select sheet
    sheet = wb["News"]

    #add color background
    sheet['A3'].fill = PatternFill(start_color="6869ee", end_color="6869ee", fill_type = "solid")
    sheet['B3'].fill = PatternFill(start_color="6869ee", end_color="6869ee", fill_type = "solid")
    sheet['C3'].fill = PatternFill(start_color="6869ee", end_color="6869ee", fill_type = "solid")

    # Assign Data
    sheet['A1'] = 'Excel creation date : ' + today_date
    sheet['B1']= 'News Crawling'
    #
    sheet['A3'] = 'Title'
    sheet['A3'].alignment = Alignment(horizontal='center')
    sheet['B3'] = 'Post Date'
    sheet['B3'].alignment = Alignment(horizontal='center')
    sheet['C3'] = 'Link'
    sheet['C3'].alignment = Alignment(horizontal='center')

    #bold text
    sheet['B1'].font = Font(bold=True, size=20)
    sheet['A3'].font = Font(bold=True)
    sheet['B3'].font = Font(bold=True)
    sheet['C3'].font = Font(bold=True)

    #coloums width
    sheet.column_dimensions['A'].width = 100
    sheet.column_dimensions['B'].width = 20    
    sheet.column_dimensions['C'].width = 30

    #LOOP BY ADD 2 COLUMNS
    for i in range(len(post_full_date)):
        sheet.cell(row=4+i, column=1).value = post_full_title[i]
        sheet.cell(row=4+i, column=2).value = post_full_date[i]
        sheet.cell(row=4+i, column=2).alignment = Alignment(horizontal='center')
        sheet.cell(row=4+i, column=3).value = '=HYPERLINK("{}", "{}")'.format(post_full_link[i], "Link")
        sheet.cell(row=4+i, column=3).font = Font(color='6869ee')
        sheet.cell(row=4+i, column=3).alignment = Alignment(horizontal='center')

def main():
    # create an empty array to store the inputs
    inputs = []
    print("Keywords: 'ABF substrate,EV,Display industry, PCB & IC substrate,Automotive industry,Semicon Industry,ASIC,Processor, Semiconductor materials ccl,laminate prices,kingsus,nan ya,unimicron'")
    # get input from the user and append to the array 3 times
    for i in range(3):
        user_input = input("Enter a keyword "+ str(i+1) +":")
        inputs.append(user_input)

    main_sub(inputs)
    print("Done!!")

if __name__ == '__main__':
    main()